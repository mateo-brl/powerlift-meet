#!/usr/bin/env python3
"""Extrait les tables de référence fédérales (barèmes publics) et des golden values
ANONYMISÉES depuis les feuilles de match .xlsm, vers des JSON consommés par le moteur
`@powerlift-meet/core` et ses tests.

Aucune donnée personnelle n'est écrite : on ne conserve ni nom, ni prénom, ni licence, ni
date de naissance complète (seule l'année est gardée, non identifiante seule).

Usage : python3 scripts/extract_reference.py "/chemin/dossier/Feuille de match"
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

AGE_CATS = ["SJR", "JR", "SNR", "M1", "M2", "M3", "M4"]
SUBCATS = ["Subjunior", "Junior", "Open", "Master1", "Master2", "Master3", "Master4"]
LEVELS_FULL = ["R3", "R2", "R1", "N2", "N1", "Europe", "Monde"]


def norm_weight_cat(raw):
    """'120 +' -> '+120 Kg' ; '53' -> '53 Kg' ; '84 +' -> '+84 Kg'."""
    s = str(raw).strip().replace(",", ".")
    s = s.replace(" +", "+").replace("+ ", "+")
    if s.endswith("+"):
        return "+" + s[:-1].strip() + " Kg"
    return s + " Kg"


def num(v):
    if v is None or v == "":
        return None
    try:
        return round(float(v), 6)
    except (TypeError, ValueError):
        return None


def extract_cat_age(wb):
    """Feuille 'Cat. d'âge' : année de naissance -> catégorie, par discipline et par année
    d'application (A = année courante, A+1 = année suivante)."""
    ws = wb["Cat. d'âge"]
    out = {"fa": {}, "pl": {}, "faA1": {}, "plA1": {}}
    # colonnes (année, catégorie) pour chaque table
    blocks = {"fa": (1, 2), "pl": (4, 5), "faA1": (7, 8), "plA1": (10, 11)}
    for key, (cy, cc) in blocks.items():
        for r in range(5, 95):
            y = ws.cell(row=r, column=cy).value
            c = ws.cell(row=r, column=cc).value
            if y is None or c is None:
                continue
            year = y.year if hasattr(y, "year") else int(y)
            out[key][str(year)] = str(c).strip()
    return out


def parse_niv_sheet(ws, has_intl):
    """Parse une feuille Niv FA / Niv PL : blocs '<AgeCat> masculin'/'féminine'.
    has_intl : la discipline a-t-elle Europe/Monde pour SJR/JR/SNR (FA & PL oui sauf masters)."""
    result = {"M": {}, "F": {}}
    maxr = ws.max_row
    # repère les en-têtes de blocs
    pat = re.compile(r"^(SJR|JR|SNR|M1|M2|M3|M4)\s+(masculin|f[ée]minine?)", re.I)
    headers = []  # (row, col, agecat, sex)
    for r in range(1, maxr + 1):
        for c in (1, 4, 10, 14):  # A, D, J, N (PL SNR est en D ; SNR fém en N)
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                m = pat.match(v.strip())
                if m:
                    sex = "M" if "mascul" in v.lower() else "F"
                    headers.append((r, c, m.group(1).upper(), sex))
    for (r, c, agecat, sex) in headers:
        # les niveaux ont Europe/Monde uniquement pour SJR/JR/SNR (8 colonnes), 6 sinon
        levels = LEVELS_FULL if (has_intl and agecat in ("SJR", "JR", "SNR")) else ["R3", "R2", "R1", "N2", "N1"]
        cats = {}
        rr = r + 2  # saute l'en-tête + la ligne 'Catég.'
        while rr <= maxr:
            cat = ws.cell(row=rr, column=c).value
            if cat is None or (isinstance(cat, str) and not re.match(r"^\s*\d", cat)):
                break
            wc = norm_weight_cat(cat)
            vals = {}
            for i, lvl in enumerate(levels):
                vals[lvl] = num(ws.cell(row=rr, column=c + 1 + i).value)
            cats[wc] = vals
            rr += 1
        result[sex][agecat] = cats
    return result


def parse_records(ws):
    """Feuille Records *_* : par catégorie de poids et sous-catégorie d'âge -> records par
    mouvement. Les noms d'athlètes (détenteurs de records publiés) sont conservés car publics."""
    out = {}
    for r in range(4, ws.max_row + 1):
        wcat = ws.cell(row=r, column=1).value  # ex '-53kg'
        sub = ws.cell(row=r, column=2).value   # ex 'Subjunior'
        if not wcat or not sub:
            continue
        wcat = str(wcat).strip()
        sub = str(sub).strip()
        entry = {}
        # blocs : Squat C/D/E, DC F/G/H, SDT I/J/K, TOTAL L/M/N
        for lift, base in (("squat", 3), ("bench", 6), ("deadlift", 9), ("total", 12)):
            athlete = ws.cell(row=r, column=base).value
            weight = num(ws.cell(row=r, column=base + 1).value)
            date = ws.cell(row=r, column=base + 2).value
            if weight is not None:
                entry[lift] = {
                    "weight": weight,
                    "athlete": str(athlete).strip() if athlete else None,
                    "context": str(date).strip() if date else None,
                }
        out.setdefault(wcat, {})[sub] = entry
    return out


# Mapping colonnes feuille FA (1-based)
COL = {"licence": 1, "club": 2, "sexe": 3, "dob": 4, "cat_age": 5, "nom": 6, "prenom": 7,
       "bodyweight": 8, "cat_poids": 9, "indice": 10, "lot": 11, "total_annonce": 21,
       "total_realise": 22, "clas": 23, "points": 24, "niv_cat": 25, "niv_open": 26,
       "discipline": 27}


def extract_golden(wb, source_id):
    """Golden values ANONYMISÉES depuis la feuille FA : entrées (sexe, année naiss., poids,
    discipline) -> sorties calculées par Excel (cat. âge, cat. poids, indice)."""
    ws = wb["FA"]
    rows = []
    idx = 0
    for r in range(9, 29):
        sexe = ws.cell(row=r, column=COL["sexe"]).value
        dob = ws.cell(row=r, column=COL["dob"]).value
        bw = ws.cell(row=r, column=COL["bodyweight"]).value
        disc = ws.cell(row=r, column=COL["discipline"]).value
        if not sexe or bw is None:
            continue
        year = dob.year if hasattr(dob, "year") else None
        rows.append({
            "id": f"{source_id}-{idx}",
            "sexe": str(sexe).strip(),
            "birthYear": year,
            "bodyweight": num(bw),
            "discipline": str(disc).strip() if disc else None,
            "expected": {
                "catAge": _s(ws.cell(row=r, column=COL["cat_age"]).value),
                "catPoids": _s(ws.cell(row=r, column=COL["cat_poids"]).value),
                "indice": num(ws.cell(row=r, column=COL["indice"]).value),
            },
        })
        idx += 1
    return rows


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def main():
    folder = Path(sys.argv[1] if len(sys.argv) > 1 else
                  "/home/matal/Documents/USOD POWER MEET/Feuille de match")
    files = sorted(folder.glob("*.xlsm"))
    if not files:
        print("Aucun .xlsm trouvé dans", folder, file=sys.stderr)
        sys.exit(1)

    repo = Path(__file__).resolve().parent.parent
    refdir = repo / "packages/core/src/reference"
    fixdir = repo / "packages/core/test/fixtures"
    refdir.mkdir(parents=True, exist_ok=True)
    fixdir.mkdir(parents=True, exist_ok=True)

    # Tables de référence : depuis le premier fichier (identiques entre plateaux)
    base = files[0]
    print("Référence depuis :", base.name)
    wb = openpyxl.load_workbook(base, data_only=True, keep_vba=True)

    cat_age = extract_cat_age(wb)
    niv_fa = parse_niv_sheet(wb["Niv FA"], has_intl=True)
    niv_pl = parse_niv_sheet(wb["Niv PL"], has_intl=True)
    records = {
        "FA": {"M": parse_records(wb["Records FA_H"]), "F": parse_records(wb["Records_FA_F"])},
        "PL": {"M": parse_records(wb["Records PL_H"]), "F": parse_records(wb["Records PL_F"])},
    }

    (refdir / "cat-age.json").write_text(json.dumps(cat_age, ensure_ascii=False, indent=2))
    (refdir / "niv-fa.json").write_text(json.dumps(niv_fa, ensure_ascii=False, indent=2))
    (refdir / "niv-pl.json").write_text(json.dumps(niv_pl, ensure_ascii=False, indent=2))
    (refdir / "records.json").write_text(json.dumps(records, ensure_ascii=False, indent=2))
    print(f"  cat-age: {sum(len(v) for v in cat_age.values())} entrées")
    print(f"  niv-fa : M={list(niv_fa['M'])} F={list(niv_fa['F'])}")
    print(f"  niv-pl : M={list(niv_pl['M'])} F={list(niv_pl['F'])}")
    print(f"  records: FA/M cats={len(records['FA']['M'])} ...")

    # Golden anonymisés depuis tous les fichiers
    golden = []
    for f in files:
        sid = re.sub(r"[^A-Za-z0-9]+", "", f.stem)[:12]
        wbf = openpyxl.load_workbook(f, data_only=True, keep_vba=True)
        if "FA" in wbf.sheetnames:
            golden.extend(extract_golden(wbf, sid))
    (fixdir / "golden-athletes.json").write_text(json.dumps(golden, ensure_ascii=False, indent=2))
    print(f"Golden : {len(golden)} athlètes anonymisés -> {fixdir/'golden-athletes.json'}")


if __name__ == "__main__":
    main()
